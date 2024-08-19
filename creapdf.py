import jinja2
import pdfkit
import openpyxl
import os
import keyring
import yagmail

from tabulate import tabulate


## extraer datos de archivo de excel
excel_dataframe = openpyxl.load_workbook("empleados.xlsx")

dataframe=excel_dataframe.active

data = []

for row in range (1, dataframe.max_row):
    _row = [row,] #subfila , agregar indice 
    
    for col in dataframe.iter_cols(1, dataframe.max_column):
        _row.append(col[row].value) #obtener valores

    data.append(_row) #agregar a la lista "data"

# tabular datos
print(tabulate(data))

## funcion para enviar correo con los pdf generados

def enviarCorreo():
    #datos inicio sesion remitente 
    email = 'pruebaspython351@gmail.com'
    password = 'jubf bitl krgl udkl'

    yag = yagmail.SMTP(user= email, password= password)

    #Contenido de Email

    id = employee[1]
    name = employee[2]+employee[3]
    destinatarios = employee[4]
    asunto = 'Diploma curso Python'
    mensaje = 'Te envio tu diploma del curso que acabas de concluir, felicidades!!!'
    html = f'<h1>Hola! {employee[2]} !!! </h1>'
    archivo = f'/home/josmacpac/Documents/Python/generadorPdf/archivos/{id} - {name}.pdf' #cambia en cada iteracion

    yag.send(destinatarios, asunto, [html, mensaje], attachments=[archivo]) 

    print('Correo enviado correctamente')
    
    ########funcion enviar correo


## Generar archivo pdf
def crea_pdf(ruta_template, info, rutacss='estilos.css'):
    nombre_template = ruta_template.split('/')[-1]
    ruta_template= ruta_template.replace(nombre_template, '')

    env = jinja2.Environment(loader=jinja2.FileSystemLoader(ruta_template))
    template = env.get_template(nombre_template)
    html = template.render(info)

    options = {'page-size': 'Letter','margin-top': '0.05in',
    'margin-right': '0.05in',
    'margin-bottom': '0.05in',
    'margin-left': '0.05in','orientation': 'Landscape','enable-local-file-access': None,'encoding': 'UTF-8'}


    config = pdfkit.configuration(wkhtmltopdf= '/usr/bin/wkhtmltopdf')
    name = employee[2]+employee[3]
    id = employee[1]
    correo = employee[4]
    ruta_salida = f'/home/josmacpac/Documents/Python/generadorPdf/archivos/{id} - {name}.pdf'
    pdfkit.from_string(html, ruta_salida, css= rutacss, options= options, configuration = config)
    print(ruta_salida)
    print(id, name, correo)


    

if __name__=="__main__":
    ruta_template = "/home/josmacpac/Documents/Python/generadorPdf/template.html"

    for employee in data:
        info = {"nombreColaborador": employee[2] + ' ' +  employee[3], "nombreCurso": "Python Basico"}
        crea_pdf(ruta_template, info)
        enviarCorreo()

    print("Los archvivos PDF's fueron creados  y enviados con exito!!!")
